import csv
from datetime import datetime
from collections import defaultdict
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment


def lese_gls_konto(pfad):
    """Liest GLS_Konto.csv und gibt Liste von (buchungstag, valutadatum, betrag, betreff) zurück."""
    buchungen = []
    with open(pfad, encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter=";")
        for row in reader:
            buchungstag = datetime.strptime(row["Buchungstag"], "%d.%m.%Y").date()
            valutadatum = datetime.strptime(row["Valutadatum"], "%d.%m.%Y").date()
            betrag_str = row["Betrag"].replace(".", "").replace(",", ".")
            betrag = round(float(betrag_str), 2)
            betreff = row["Verwendungszweck"]
            buchungen.append((buchungstag, valutadatum, betrag, betreff))
    return buchungen


def lese_gls_buchhaltung(pfad):
    """Liest GLS_Buchhaltung.xlsx und gibt Liste von (datum, betrag, betreff) zurück."""
    buchungen = []
    wb = load_workbook(pfad)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    idx_datum = headers.index("Datum")
    idx_text = headers.index("Buchungstext")
    idx_soll = headers.index("Gutschrift / Soll")
    idx_haben = headers.index("Lastschrift / Haben")

    for row in ws.iter_rows(min_row=2, values_only=True):
        datum_val = row[idx_datum]
        if datum_val is None:
            continue  # z.B. Anfangssaldo-Zeile
        if isinstance(datum_val, datetime):
            datum = datum_val.date()
        else:
            continue

        soll = row[idx_soll]  # Gutschrift = positiv
        haben = row[idx_haben]  # Lastschrift = negativ

        if soll is not None:
            betrag = round(float(soll), 2)
        elif haben is not None:
            betrag = round(-float(haben), 2)
        else:
            continue

        betreff = row[idx_text] or ""
        buchungen.append((datum, betrag, betreff))
    return buchungen


def abgleich(gls_buchungen, bh_buchungen):
    """Matcht Buchungen in zwei Schritten. Gibt drei Listen zurück.

    Schritt 1: Match auf (Buchungstag == Buchung Buchhaltung) und Betrag.
    Schritt 2: Vom Rest Match auf (Valutadatum == Buchung Buchhaltung) und Betrag.
    """
    uebereinstimmend = []

    # -- Schritt 1: Buchungstag == Datum Buchhaltung --
    gls_map1 = defaultdict(list)
    for i, (buchungstag, valutadatum, betrag, betreff) in enumerate(gls_buchungen):
        gls_map1[(buchungstag, betrag)].append(i)

    bh_map1 = defaultdict(list)
    for j, (datum_bh, betrag, betreff) in enumerate(bh_buchungen):
        bh_map1[(datum_bh, betrag)].append(j)

    matched_gls = set()
    matched_bh = set()

    for key in gls_map1:
        if key in bh_map1:
            gls_idx = gls_map1[key]
            bh_idx = bh_map1[key]
            for k in range(min(len(gls_idx), len(bh_idx))):
                gi, bi = gls_idx[k], bh_idx[k]
                buchungstag, valutadatum, betrag, betreff_gls = gls_buchungen[gi]
                datum_bh, _, betreff_bh = bh_buchungen[bi]
                uebereinstimmend.append(
                    (datum_bh, buchungstag, valutadatum, betrag, betreff_gls, betreff_bh)
                )
                matched_gls.add(gi)
                matched_bh.add(bi)

    # -- Schritt 2: Valutadatum == Datum Buchhaltung (nur übrige) --
    gls_rest = [(i, gls_buchungen[i]) for i in range(len(gls_buchungen)) if i not in matched_gls]
    bh_rest = [(j, bh_buchungen[j]) for j in range(len(bh_buchungen)) if j not in matched_bh]

    gls_map2 = defaultdict(list)
    for idx, (_, (buchungstag, valutadatum, betrag, betreff)) in enumerate(gls_rest):
        gls_map2[(valutadatum, betrag)].append(idx)

    bh_map2 = defaultdict(list)
    for idx, (_, (datum_bh, betrag, betreff)) in enumerate(bh_rest):
        bh_map2[(datum_bh, betrag)].append(idx)

    matched_gls2 = set()
    matched_bh2 = set()

    for key in gls_map2:
        if key in bh_map2:
            gls_idx = gls_map2[key]
            bh_idx = bh_map2[key]
            for k in range(min(len(gls_idx), len(bh_idx))):
                gi, bi = gls_idx[k], bh_idx[k]
                _, (buchungstag, valutadatum, betrag, betreff_gls) = gls_rest[gi]
                _, (datum_bh, _, betreff_bh) = bh_rest[bi]
                uebereinstimmend.append(
                    (datum_bh, buchungstag, valutadatum, betrag, betreff_gls, betreff_bh)
                )
                matched_gls2.add(gi)
                matched_bh2.add(bi)

    # -- Verbleibender Rest --
    nur_gls = []
    for idx, (_, (buchungstag, valutadatum, betrag, betreff)) in enumerate(gls_rest):
        if idx not in matched_gls2:
            nur_gls.append((None, buchungstag, valutadatum, betrag, betreff, ""))

    nur_bh = []
    for idx, (_, (datum_bh, betrag, betreff)) in enumerate(bh_rest):
        if idx not in matched_bh2:
            nur_bh.append((datum_bh, None, None, betrag, "", betreff))

    uebereinstimmend.sort(key=lambda x: (x[0] or x[1], x[3]))
    nur_gls.sort(key=lambda x: (x[1], x[3]))
    nur_bh.sort(key=lambda x: (x[0], x[3]))

    return nur_gls, nur_bh, uebereinstimmend


def schreibe_ergebnis(pfad, nur_gls, nur_bh, uebereinstimmend):
    """Schreibt das Ergebnis in eine xlsx-Datei."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Kontoabgleich GLS"

    # Header
    headers = [
        "Buchung (Buchhaltung)", "Buchungstag", "Valutadatum",
        "Betrag", "Betreff GLS", "Betreff Buchhaltung", "Status",
    ]
    ws.append(headers)

    # Header-Formatierung
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Farben für Status
    fill_nur_gls = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    fill_nur_bh = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    fill_ok = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

    def schreibe_block(daten, status, fill):
        for datum_bh, buchungstag, valutadatum, betrag, betreff_gls, betreff_bh in daten:
            ws.append([datum_bh, buchungstag, valutadatum, betrag, betreff_gls, betreff_bh, status])
            for cell in ws[ws.max_row]:
                cell.fill = fill

    schreibe_block(nur_gls, "nur GLS", fill_nur_gls)
    schreibe_block(nur_bh, "nur Buchhaltung", fill_nur_bh)
    schreibe_block(uebereinstimmend, "übereinstimmend", fill_ok)

    # Spaltenbreiten
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 60
    ws.column_dimensions["F"].width = 60
    ws.column_dimensions["G"].width = 20

    # Betrag als Zahl formatieren
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = '#,##0.00'

    # Datum formatieren
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=3):
        for cell in row:
            cell.number_format = 'DD.MM.YYYY'

    wb.save(pfad)


def main():
    gls_buchungen = lese_gls_konto("kontoabgleich/GLS_Konto.csv")
    bh_buchungen = lese_gls_buchhaltung("kontoabgleich/GLS_Buchhaltung.xlsx")

    print(f"GLS Konto: {len(gls_buchungen)} Buchungen")
    print(f"Buchhaltung: {len(bh_buchungen)} Buchungen")

    nur_gls, nur_bh, uebereinstimmend = abgleich(gls_buchungen, bh_buchungen)

    print(f"\nErgebnis:")
    print(f"  Übereinstimmend:  {len(uebereinstimmend)}")
    print(f"  Nur GLS:          {len(nur_gls)}")
    print(f"  Nur Buchhaltung:  {len(nur_bh)}")

    schreibe_ergebnis("kontoabgleich_gls.xlsx", nur_gls, nur_bh, uebereinstimmend)
    print("\nDatei geschrieben: kontoabgleich_gls.xlsx")


if __name__ == "__main__":
    main()
