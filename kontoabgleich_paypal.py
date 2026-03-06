import csv
from datetime import datetime
from collections import defaultdict
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment


def lese_paypal_konto(pfad):
    """Liest Paypal_Konto.csv und gibt Liste von (datum, betrag, betreff) zurück."""
    buchungen = []
    with open(pfad, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            datum = datetime.strptime(row["Datum"], "%d.%m.%Y").date()
            betrag_str = row["Brutto"].replace(".", "").replace(",", ".")
            betrag = round(float(betrag_str), 2)
            name = row["Name"].strip()
            hinweis = row["Hinweis"].strip()
            typ = row["Typ"].strip()
            betreff = f"{name} | {hinweis}" if hinweis else f"{name} ({typ})"
            
            buchungen.append((datum, betrag, betreff))
    return buchungen


def lese_paypal_buchhaltung(pfad):
    """Liest Paypal_Buchhaltung.xlsx und gibt Liste von (datum, betrag, betreff) zurück."""
    buchungen = []
    wb = load_workbook(pfad)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    idx_datum = headers.index("Datum")
    idx_text = headers.index("Buchungstext")
    idx_soll = headers.index("Gutschrift / Soll")
    idx_haben = headers.index("Lastschrift / Haben")

    for row in ws.iter_rows(min_row=2, values_only=True):
        datum_val = row[idx_datum]
        if datum_val is None:
            continue
        if isinstance(datum_val, datetime):
            datum = datum_val.date()
        else:
            continue

        soll = row[idx_soll]
        haben = row[idx_haben]

        if soll is not None:
            betrag = round(float(soll), 2)
        elif haben is not None:
            betrag = round(-float(haben), 2)
        else:
            continue

        betreff = row[idx_text] or ""
        buchungen.append((datum, betrag, betreff))
    return buchungen


def abgleich(pp_buchungen, bh_buchungen):
    """Matcht Buchungen anhand (Datum, Betrag). Gibt drei Listen zurück."""

    pp_map = defaultdict(list)
    for datum, betrag, betreff in pp_buchungen:
        pp_map[(datum, betrag)].append(betreff)

    bh_map = defaultdict(list)
    for datum, betrag, betreff in bh_buchungen:
        bh_map[(datum, betrag)].append(betreff)

    alle_keys = set(pp_map.keys()) | set(bh_map.keys())

    uebereinstimmend = []
    nur_pp = []
    nur_bh = []

    for key in sorted(alle_keys):
        datum, betrag = key
        pp_liste = pp_map.get(key, [])
        bh_liste = bh_map.get(key, [])

        anzahl_match = min(len(pp_liste), len(bh_liste))

        for i in range(anzahl_match):
            uebereinstimmend.append((datum, betrag, pp_liste[i], bh_liste[i]))

        for i in range(anzahl_match, len(pp_liste)):
            nur_pp.append((datum, betrag, pp_liste[i], ""))

        for i in range(anzahl_match, len(bh_liste)):
            nur_bh.append((datum, betrag, "", bh_liste[i]))

    return nur_pp, nur_bh, uebereinstimmend


def schreibe_ergebnis(pfad, nur_pp, nur_bh, uebereinstimmend):
    """Schreibt das Ergebnis in eine xlsx-Datei."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Kontoabgleich PayPal"

    headers = ["Datum", "Betrag", "Betreff PayPal", "Betreff Buchhaltung", "Status"]
    ws.append(headers)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    fill_nur_pp = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    fill_nur_bh = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    fill_ok = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

    def schreibe_block(daten, status, fill):
        for datum, betrag, betreff_pp, betreff_bh in daten:
            ws.append([datum, betrag, betreff_pp, betreff_bh, status])
            for cell in ws[ws.max_row]:
                cell.fill = fill

    schreibe_block(nur_pp, "nur PayPal", fill_nur_pp)
    schreibe_block(nur_bh, "nur Buchhaltung", fill_nur_bh)
    schreibe_block(uebereinstimmend, "übereinstimmend", fill_ok)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 20

    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = '#,##0.00'

    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            cell.number_format = 'DD.MM.YYYY'

    wb.save(pfad)


def main():
    pp_buchungen = lese_paypal_konto("kontoabgleich/Paypal_Konto.csv")
    bh_buchungen = lese_paypal_buchhaltung("kontoabgleich/Paypal_Buchhaltung.xlsx")

    print(f"PayPal Konto: {len(pp_buchungen)} Buchungen")
    print(f"Buchhaltung:  {len(bh_buchungen)} Buchungen")

    nur_pp, nur_bh, uebereinstimmend = abgleich(pp_buchungen, bh_buchungen)

    print(f"\nErgebnis:")
    print(f"  Übereinstimmend:  {len(uebereinstimmend)}")
    print(f"  Nur PayPal:       {len(nur_pp)}")
    print(f"  Nur Buchhaltung:  {len(nur_bh)}")

    schreibe_ergebnis("kontoabgleich_paypal.xlsx", nur_pp, nur_bh, uebereinstimmend)
    print("\nDatei geschrieben: kontoabgleich_paypal.xlsx")


if __name__ == "__main__":
    main()
